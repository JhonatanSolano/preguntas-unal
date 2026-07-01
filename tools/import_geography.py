import argparse
import ast
import json
import re
import time
import urllib.parse
import urllib.request
from pathlib import Path

from cryptography.hazmat.primitives import hashes, serialization
from cryptography.hazmat.primitives.asymmetric import padding
from openpyxl import load_workbook


TOKEN_URL = "https://oauth2.googleapis.com/token"
FIRESTORE_SCOPE = "https://www.googleapis.com/auth/datastore"


def b64url(data: bytes) -> str:
    import base64
    return base64.urlsafe_b64encode(data).rstrip(b"=").decode("ascii")


def get_access_token(service_account: dict) -> str:
    now = int(time.time())
    header = {"alg": "RS256", "typ": "JWT"}
    payload = {
        "iss": service_account["client_email"],
        "scope": FIRESTORE_SCOPE,
        "aud": TOKEN_URL,
        "iat": now,
        "exp": now + 3600,
    }
    unsigned = f"{b64url(json.dumps(header, separators=(',', ':')).encode())}.{b64url(json.dumps(payload, separators=(',', ':')).encode())}"
    private_key = serialization.load_pem_private_key(service_account["private_key"].encode(), password=None)
    signature = private_key.sign(unsigned.encode(), padding.PKCS1v15(), hashes.SHA256())
    assertion = f"{unsigned}.{b64url(signature)}"
    body = urllib.parse.urlencode({
        "grant_type": "urn:ietf:params:oauth:grant-type:jwt-bearer",
        "assertion": assertion,
    }).encode()
    request = urllib.request.Request(TOKEN_URL, data=body, method="POST")
    request.add_header("Content-Type", "application/x-www-form-urlencoded")
    with urllib.request.urlopen(request, timeout=30) as response:
        return json.loads(response.read().decode())["access_token"]


def fs_value(value):
    if value is None:
        return {"nullValue": None}
    if isinstance(value, bool):
        return {"booleanValue": value}
    if isinstance(value, int):
        return {"integerValue": str(value)}
    if isinstance(value, float):
        return {"doubleValue": value}
    return {"stringValue": str(value)}


def fs_fields(data: dict) -> dict:
    return {key: fs_value(value) for key, value in data.items()}


def doc_name(project_id: str, relative_path: str) -> str:
    parts = [urllib.parse.quote(part, safe="") for part in relative_path.strip("/").split("/")]
    return f"projects/{project_id}/databases/(default)/documents/{'/'.join(parts)}"


def title_es(text: str) -> str:
    text = " ".join(str(text or "").strip().split())
    if not text:
        return ""
    small = {"de", "del", "la", "las", "los", "y", "e", "el"}
    words = []
    for idx, word in enumerate(text.lower().split(" ")):
        if idx > 0 and word in small:
            words.append(word)
        else:
            words.append("-".join(part.capitalize() for part in word.split("-")))
    result = " ".join(words)
    replacements = {
        "D.c.": "D.C.",
        "Dc": "D.C.",
        "San Andres": "San Andrés",
        "Providencia y Santa Catalina": "Providencia y Santa Catalina",
    }
    for old, new in replacements.items():
        result = result.replace(old, new)
    return result


def read_colombia_divipola(path: Path):
    workbook = load_workbook(path, read_only=True, data_only=True)
    worksheet = workbook["Municipios"] if "Municipios" in workbook.sheetnames else workbook[workbook.sheetnames[0]]
    departments = {}
    municipalities = []

    for row in worksheet.iter_rows(min_row=12, values_only=True):
        department_code, department_name, municipality_code, municipality_name = row[:4]
        if not department_code or not municipality_code:
            continue
        department_code = str(department_code).strip().zfill(2)
        municipality_code = str(municipality_code).strip().zfill(5)
        department_official = str(department_name or "").strip()
        municipality_official = str(municipality_name or "").strip()
        if not department_official or not municipality_official:
            continue
        departments[department_code] = {
            "id": department_code,
            "countryId": "CO",
            "code": department_code,
            "name": title_es(department_official),
            "officialName": department_official,
            "kind": "department",
        }
        municipalities.append({
            "id": municipality_code,
            "countryId": "CO",
            "regionId": department_code,
            "departmentCode": department_code,
            "departmentName": title_es(department_official),
            "departmentOfficialName": department_official,
            "code": municipality_code,
            "municipalityCode": municipality_code,
            "name": title_es(municipality_official),
            "municipalityName": title_es(municipality_official),
            "officialName": municipality_official,
        })

    return sorted(departments.values(), key=lambda item: item["name"]), sorted(municipalities, key=lambda item: (item["departmentName"], item["name"]))


def extract_sql_insert_rows(sql: str, table: str):
    pattern = rf"INSERT\s+INTO\s+`{re.escape(table)}`\s*\([^)]*\)\s*VALUES\s*(.*?);"
    match = re.search(pattern, sql, re.IGNORECASE | re.DOTALL)
    if not match:
        return []
    values = match.group(1).strip()
    rows = re.findall(r"\((.*?)\)(?:,|$)", values, re.DOTALL)
    parsed = []
    for row in rows:
        clean = row.replace("\\'", "'")
        parsed.append(ast.literal_eval(f"({clean})"))
    return parsed


def read_venezuela_sql(path: Path):
    sql = path.read_text(encoding="utf-8")
    states_raw = extract_sql_insert_rows(sql, "estados")
    municipalities_raw = extract_sql_insert_rows(sql, "municipios")
    states_by_id = {}
    regions = []
    municipalities = []

    for state_id, state_name, iso in states_raw:
        code = str(iso or f"VE-{state_id}").strip()
        region_id = str(state_id).zfill(2)
        state = {
            "id": region_id,
            "countryId": "VE",
            "code": code,
            "numericCode": region_id,
            "name": str(state_name).strip(),
            "officialName": str(state_name).strip(),
            "kind": "state",
            "source": "github.com/marydn/venezuela-sql",
        }
        states_by_id[int(state_id)] = state
        regions.append(state)

    for municipality_id, state_id, municipality_name in municipalities_raw:
        state = states_by_id.get(int(state_id))
        if not state:
            continue
        municipality_code = str(municipality_id).zfill(3)
        municipalities.append({
            "id": municipality_code,
            "countryId": "VE",
            "regionId": state["id"],
            "stateCode": state["code"],
            "stateName": state["name"],
            "code": municipality_code,
            "municipalityCode": municipality_code,
            "name": str(municipality_name).strip(),
            "municipalityName": str(municipality_name).strip(),
            "officialName": str(municipality_name).strip(),
            "source": "github.com/marydn/venezuela-sql",
        })

    return sorted(regions, key=lambda item: item["name"]), sorted(municipalities, key=lambda item: (item["stateName"], item["name"]))


def build_writes(project_id: str, departments, municipalities, venezuela_regions=None, venezuela_municipalities=None):
    venezuela_regions = venezuela_regions or []
    venezuela_municipalities = venezuela_municipalities or []
    countries = [
        {"id": "CO", "name": "Colombia", "iso2": "CO", "iso3": "COL", "phoneCode": "+57", "regionLabel": "Departamento", "municipalityLabel": "Municipio"},
        {"id": "VE", "name": "Venezuela", "iso2": "VE", "iso3": "VEN", "phoneCode": "+58", "regionLabel": "Estado", "municipalityLabel": "Municipio", "status": "active" if venezuela_regions else "pending_catalog"},
    ]
    writes = []
    for country in countries:
        country_id = country["id"]
        data = {key: value for key, value in country.items() if key != "id"}
        writes.append({"update": {"name": doc_name(project_id, f"countries/{country_id}"), "fields": fs_fields(data)}})

    for department in departments:
        department_id = department["id"]
        data = {key: value for key, value in department.items() if key != "id"}
        writes.append({"update": {"name": doc_name(project_id, f"countries/CO/regions/{department_id}"), "fields": fs_fields(data)}})

    for municipality in municipalities:
        municipality_id = municipality["id"]
        region_id = municipality["regionId"]
        data = {key: value for key, value in municipality.items() if key != "id"}
        writes.append({"update": {"name": doc_name(project_id, f"countries/CO/regions/{region_id}/municipalities/{municipality_id}"), "fields": fs_fields(data)}})

    for region in venezuela_regions:
        region_id = region["id"]
        data = {key: value for key, value in region.items() if key != "id"}
        writes.append({"update": {"name": doc_name(project_id, f"countries/VE/regions/{region_id}"), "fields": fs_fields(data)}})

    for municipality in venezuela_municipalities:
        municipality_id = municipality["id"]
        region_id = municipality["regionId"]
        data = {key: value for key, value in municipality.items() if key != "id"}
        writes.append({"update": {"name": doc_name(project_id, f"countries/VE/regions/{region_id}/municipalities/{municipality_id}"), "fields": fs_fields(data)}})
    return writes


def batch_write(project_id: str, token: str, writes):
    url = f"https://firestore.googleapis.com/v1/projects/{project_id}/databases/(default)/documents:batchWrite"
    total = len(writes)
    for start in range(0, total, 450):
        chunk = writes[start:start + 450]
        request = urllib.request.Request(url, data=json.dumps({"writes": chunk}).encode(), method="POST")
        request.add_header("Authorization", f"Bearer {token}")
        request.add_header("Content-Type", "application/json")
        with urllib.request.urlopen(request, timeout=60) as response:
            result = json.loads(response.read().decode())
        errors = [item.get("status") for item in result.get("writeResults", []) if item.get("status")]
        if errors:
            raise RuntimeError(f"Firestore devolvió errores en el lote {start // 450 + 1}: {errors[:3]}")
        print(f"Subidos {min(start + len(chunk), total)} de {total} documentos...")


def main():
    parser = argparse.ArgumentParser(description="Importa catálogo geográfico a Cloud Firestore.")
    parser.add_argument("--service-account", required=True, help="Ruta del JSON privado de Firebase Admin SDK.")
    parser.add_argument("--colombia-divipola", required=True, help="Ruta del Excel DIVIPOLA_Municipios.xlsx del DANE.")
    parser.add_argument("--venezuela-sql", help="Ruta del SQL de marydn/venezuela-sql con estados y municipios.")
    parser.add_argument("--dry-run", action="store_true", help="Solo lee y muestra conteos; no escribe en Firestore.")
    args = parser.parse_args()

    service_account_path = Path(args.service_account)
    divipola_path = Path(args.colombia_divipola)
    service_account = json.loads(service_account_path.read_text(encoding="utf-8"))
    project_id = service_account["project_id"]

    departments, municipalities = read_colombia_divipola(divipola_path)
    venezuela_regions = []
    venezuela_municipalities = []
    if args.venezuela_sql:
        venezuela_regions, venezuela_municipalities = read_venezuela_sql(Path(args.venezuela_sql))
    print(f"Proyecto: {project_id}")
    print(f"Colombia: {len(departments)} departamentos y {len(municipalities)} municipios/áreas leídos.")
    if args.venezuela_sql:
        print(f"Venezuela: {len(venezuela_regions)} estados/distritos y {len(venezuela_municipalities)} municipios leídos.")
    else:
        print("Venezuela: se crea el país VE, pero los estados/municipios quedan pendientes hasta tener fuente confiable.")

    if args.dry_run:
        print("Dry run completado. No se escribió en Firestore.")
        return

    writes = build_writes(project_id, departments, municipalities, venezuela_regions, venezuela_municipalities)
    token = get_access_token(service_account)
    batch_write(project_id, token, writes)
    print("Importación terminada.")


if __name__ == "__main__":
    main()
